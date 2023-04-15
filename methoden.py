from selenium import webdriver
from selenium.webdriver.common.keys import Keys
import time
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from datetime import datetime
from selenium import webdriver
from webdriver_manager.chrome import ChromeDriverManager
import os

def haupt (gestern):
    d = gestern.day
    m = gestern.month
    j = gestern.year
    tag=["Montag","Dienstag","Mittwoch","Donnerstag","Freitag","Samstag","Sonntag"]
    driver = webdriver.Chrome(ChromeDriverManager().install())
    driver.get("http://desktop-7a7h4du:90/")
    search = driver.find_element(value="username")
    search.send_keys("admin")
    pwd = driver.find_element(value="dijit_form_TextBox_0")
    pwd.send_keys("kasse")
    pwd.send_keys(Keys.RETURN)
    gestern.weekday()
    zimmer = []
    zimmer=zimmmer(d,m,j,zimmer,[],20,driver)
    restzimmer= zimmmer(d,m,j,zimmer,[],51,driver)
    debitor= zimmmer(d,m,j,[],[],80,driver)

    zimmer=restzimmer[1]
    restaurant=restzimmer[0]
    debitorrest =debitor[0]
    debitorzimmer=debitor[1]
    gutscheinkauf,gutscheinnr = Gutschein(d,m,j,41,driver)
    gutscheineinloesung,gutscheinnr2= Gutschein(d,m,j,42,driver)
    for i in gutscheinnr2:
        gutscheinnr.append(i)
    zimmer=pruefen(zimmer,gutscheinnr)
    restaurant= pruefen(restaurant,gutscheinnr)
    debitorrest= pruefen(debitor[0],gutscheinnr)
    debitorzimmer= pruefen(debitor[1],gutscheinnr)
    wb= openpyxl.load_workbook("C:/Users/Gasthof Lamm/Documents/test kasse/1.xlsx")
    sheet= wb.active
    zeile =19
    tupel=[zeile,sheet]
    tupel=einfuegen(tupel,restaurant,"Restaurant")
    tupel=einfuegen(tupel,debitorrest,"")
    tupel=einfuegen(tupel,gutscheineinloesung,"Gutscheineinlösung")
    tupel=einfuegen(tupel,gutscheinkauf,"Gutscheinkauf")
    name=zimmernr(zimmer)
    tupel=einfuegen(tupel,zimmer,name)
    name=zimmernr(debitorzimmer)
    tupel=einfuegen(tupel,debitorzimmer,name)
    sheet=tupel[1]
    sheet= trinkgeld(d,m,j,driver,sheet)
    zeile =tupel[0]
    sheet=bar(sheet,gestern)
    sheet.cell(1,1,f"{tag[gestern.weekday()]} {d}.{m}.{j}")
    if not os.path.exists(f"Z:/Kassenblatt/{j}"):
        os.mkdir(f"Z:/Kassenblatt/{j}")
    if not os.path.exists(f"Z:/Kassenblatt/{j}/{m}"):
        os.mkdir(f"Z:/Kassenblatt/{j}/{m}")
    wb.save(f"Z:/Kassenblatt/{j}/{m}/{d}.xlsx")

def zimmmer(d, m, j, zimmer, restaurant, akt, driver):
    url = f'http://desktop-7a7h4du:90/journal/list/filter/%7B"businessdate":"{d}.{m}.{j}","interval":"0","zeit":"0","zcount":"0","kellner":"0","terminal":"0","site":"0","tisch":"0","artikel":"0","summe":"0","aktion":"{akt}"%7D'
    driver.get(url)
    laenge = len(driver.find_elements(By.XPATH, "//table[@width='98%']/tbody/tr/td[2]"))
    if laenge:
        laenge = laenge + 2
        for i in range(laenge):
            liste = []
            if i < 2:
                continue
            if akt == 20:
                if driver.find_element(By.XPATH,f"//table[@width='98%']/tbody/tr[{i}]/td[6]").text.isnumeric() and len(
                    driver.find_element(By.XPATH, f"//table[@width='98%']/tbody/tr[{i}]/td[6]").text) > 2:
                    for y in range(1, 12):
                        liste.append(
                            driver.find_element(By.XPATH, f"//table[@width='98%']/tbody/tr[{i}]/td[{y}]").text)
                    zimmer.append(liste)
            if akt == 51 or akt == 80:
                for y in range(1, 12):
                    liste.append(driver.find_element(By.XPATH, f"//table[@width='98%']/tbody/tr[{i}]/td[{y}]").text)
                if driver.find_element(By.XPATH,f"//table[@width='98%']/tbody/tr[{i}]/td[6]").text.isnumeric() and len(driver.find_element(By.XPATH, f"//table[@width='98%']/tbody/tr[{i}]/td[6]").text) > 2 and int(driver.find_element(By.XPATH, f"//table[@width='98%']/tbody/tr[{i}]/td[6]").text)<111 :
                    zimmer.append(liste)
                else:
                    restaurant.append(liste)

    if akt == 20:
        return zimmer
    if akt == 51:
        return (restaurant, zimmer)
    if akt == 80:
        return (restaurant, zimmer)

def Gutschein(d, m, j, akt, driver):
    url = f'http://desktop-7a7h4du:90/journal/list/filter/%7B"businessdate":"{d}.{m}.{j}","interval":"0","zeit":"0","zcount":"0","kellner":"0","terminal":"0","site":"0","tisch":"0","artikel":"0","summe":"0","aktion":"{akt}"%7D'
    driver.get(url)
    laenge = len(driver.find_elements(By.XPATH, "//table[@width='98%']/tbody/tr/td[2]"))
    gutschein = []
    gutscheinnr = []
    if laenge:
        laenge = laenge + 2
        for i in range(laenge):
            liste = []
            if i < 2:
                continue
            for y in range(1, 12):
                liste.append(driver.find_element(By.XPATH, f"//table[@width='98%']/tbody/tr[{i}]/td[{y}]").text)
            gutschein.append(liste)
    if gutschein != []:
        driver.get("http://desktop-7a7h4du:90/report/prev/von/01.07.2022/bis/25.07.2022/interval/5/kind/reports_listen_gutscheine/reltime/Dieser%20Monat/filterart//filtertime//filterkellner//filtertisch//filtersite//filtersituation//filterfs//filterweekday/")
        nummer2 = driver.find_elements(By.XPATH, "//table[@width='100%']/tbody/tr/td[4]")
        for i in gutschein:
            for k in nummer2:
                abs = i[0].split(".")
                abx = abs[0] + "." + abs[1] + "." + "20" + abs[2]
                if abx.__contains__(k.text):
                    nr = driver.find_element(By.XPATH,f"//table[@width='100%']/tbody/tr[{nummer2.index(k)+2}]/td[1]").text
                    name = f'GutscheinNr: {nr}'
                    gutschein[gutschein.index(i)].append(name)
                    break
        url = f'http://desktop-7a7h4du:90/journal/list/filter/%7B"businessdate":"{d}.{m}.{j}","interval":"0","zeit":"0","zcount":"0","kellner":"0","terminal":"0","site":"0","tisch":"{gutschein[gutschein.index(i)][5]}","artikel":"0","summe":"0","aktion":"0"%7D'
        driver.get(url)
        laenge =int(len(driver.find_elements(By.XPATH, "//table[@width='98%']/tbody/tr/td[2]")))
        liste=[]
        iteration=[]
        laenge = laenge + 2
        x=0

        iteration=gutschein.copy()
        for i in iteration:
            for y in range(laenge):
                if y < 2:
                    continue
                if driver.find_element(By.XPATH,f"//table[@width='98%']/tbody/tr[{y}]/td[7]").text == gutschein[gutschein.index(i)][6]:
                    for a in range(1, 12):
                        liste.append(driver.find_element(By.XPATH,f"//table[@width='98%']/tbody/tr[{y}]/td[{a}]").text)
                    gutscheinnr.append(liste[6])
                    if len(gutschein[gutschein.index(i)][5])>2 and gutschein[gutschein.index(i)][5].isnumeric():
                        liste.append(f"Zi Nr:{int(gutschein[gutschein.index(i)][5])-100}")
                    else:
                        liste.append("Restaurant")
                    gutschein.insert(gutschein.index(i)+1,liste)
                    break
    return gutschein,gutscheinnr







def einfuegen(tupel, liste, name):
    if liste:
        namenliste = 0
        if isinstance(name, list):
            namenliste = name
        for i in range(len(liste)):
            if isinstance(namenliste, list):
                name = f"ZimmerNr: {namenliste[i]}"
            tupel[1].cell(tupel[0], 1, name)
            tupel[1].cell(tupel[0], 2, liste[i][6])
            tupel[1].cell(tupel[0], 3, liste[i][8].replace("-", ""))
            tupel[1].cell(tupel[0], 4, liste[i][10])
            if name == "Gutscheineinlösung"and len(liste[i])<11:
                tupel[1].cell(tupel[0], 1, liste[i][11])
            tupel[0] = tupel[0] + 1

    return tupel

def pruefen(restaurant,gutscheinnr):
    for i in restaurant:
        if i[10] == 'Kartenzahlung (Gutschrift)' or i[10] == "Bar gegeben (Gutschrift)" or i[10] == "Debitorenrechnung (Gutschrift)" or i[6]in gutscheinnr:
            betrag = f"-{i[8]}"
            tisch = i[5]
            zeit = datetime.strptime(i[0],"%d.%m.%y %H:%M:%S")
            restaurant.remove(i)
            for k in restaurant:
                vergleich=datetime.strptime(k[0],"%d.%m.%y %H:%M:%S")
                if k[8] == betrag and k[5] == tisch and vergleich<=zeit:
                    restaurant.remove(k)
                    break
    return restaurant

def zimmernr(zimmer):
    name = []
    for i in range(len(zimmer)):
        das = int(zimmer[i][5]) - 100
        name.append(das)
    return name

def trinkgeld(d, m, j, driver, sheet):
    url = f'http://desktop-7a7h4du:90/journal/list/filter/%7B"businessdate":"{d}.{m}.{j}","interval":"0","zeit":"0","zcount":"0","kellner":"0","terminal":"0","site":"0","tisch":"0","artikel":"0","summe":"0","aktion":"27"%7D'
    driver.get(url)
    laenge = len(driver.find_elements(By.XPATH, "//table[@width='98%']/tbody/tr/td[2]"))
    tip = 0.0
    if laenge:
        laenge = laenge + 2
        for i in range(laenge):
            gut=0
            if i < 2:
                continue
            k = driver.find_element(By.XPATH, f"//table[@width='98%']/tbody/tr[{i}]/td[9]").text
            if driver.find_element(By.XPATH, f"//table[@width='98%']/tbody/tr[{i}]/td[11]").text== "Trinkgeld (Gutschrift)":
                gut =1
            k = k.replace(",", ".")
            k = float(k)
            if k > 0 or gut==1:
                tip = tip + k
        tip=tip.__round__(3)
        sheet.cell(6, 1, "Tip")
        sheet.cell(6, 3, f"{tip}€")
        sheet.cell(6, 4, f"-{tip}€")

    return sheet
def bar(sheet,gestern):
    driver = webdriver.Chrome(ChromeDriverManager().install())
    driver.get("http://desktop-7a7h4du:8080/amadeus/cockpit?")
    search = driver.find_element(value="username")
    search.send_keys("admin")
    pwd = driver.find_element(By.ID, "password")
    pwd.send_keys("kasse")
    pwd.send_keys(Keys.RETURN)
    driver.get(f"http://desktop-7a7h4du:8080/amadeus/cockpit?app=info&type=quickview&action=show&dateFrom=CUSTOMSPAN&start={gestern.day}.{gestern.month}.{gestern.year}&end={gestern.day}.{gestern.month}.{gestern.year}&waiter=-1")
    try:
        if driver.find_element(By.XPATH, '//*[@id="quickview"]/div[1]/div[7]/div[2]').text.__contains__("Bargeld"):
            c=driver.find_element(By.XPATH, '//*[@id="quickview"]/div[1]/div[7]/div[3]').text.replace("-","")
            c = c.replace(",", ".")
        else:
            c="0.00"
    except:
        c="0,0"
    c=c.replace(",",".")
    for i in range (len(c)-1):
        if c[i]=="." and i!=len(c)-3:
            c=c[:i]+c[i+1:]
    c=float(c)
    sheet.cell(4,3,f"{c}€")
    book = openpyxl.load_workbook('Z:/Auswertungen Kasse/2022-04-04 Zählprotokoll - neu.xlsm', keep_vba=True)
    blatt = book.get_sheet_by_name("Kasse")
    a=3
    while blatt.cell(a,1).value is not None and blatt.cell(a,1).value!=gestern :
        a=a+1
    if  blatt.cell(a,1).value!=gestern:
        while blatt.cell(a,2).value is not None:
            a+=1
    blatt.cell(a,2,c).number_format = '#,##0.00€'
    breite= 3
    zeile =9
    while breite<=5:
        if blatt.cell(a,breite).value is not None :
            sheet.cell(zeile,4,blatt.cell(a,breite).value)
            zeile+=1
        breite+=1
    betrag = iterieren(blatt.cell(a,6).value,blatt)
    betrag = betrag.__round__(3)
    sheet.cell(14,5,f"{betrag}€")
    book.save('Z:/Auswertungen Kasse/2022-04-04 Zählprotokoll - neu.xlsm')
    return sheet
def iterieren(zelle ,sheet):
    if  zelle is None:
        return 0.0
    alph=["A","B","C","D","E","F"]
    betrag=0.0
    if zelle.__class__ is float or zelle.__class__ is int :
        if zelle is not None:
            return zelle
    if zelle[0]=="=":
        if any(c.isalpha() for c in zelle):
            for i in range(len(zelle)):
                if zelle[i]  in alph:
                    spalte = alph.index(zelle[i])+1
                    zeile=""
                    while zelle[i+1]!=")"and zelle[i+1]!="+"  :
                            i=i+1
                            zeile = zeile+zelle[i]
                            if i+1>= len(zelle):
                                break
                    betragn =iterieren(sheet.cell(int(zeile), spalte).value, sheet)
                    betrag = betrag+ betragn
        else:
            return eval(zelle.replace("=",""))
        return betrag